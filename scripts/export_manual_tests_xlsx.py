# -*- coding: utf-8 -*-
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


TEST_CASES = [
    {
        "title": "Регистрация нового пользователя",
        "precondition": "Пользователь не авторизован.",
        "steps": [
            "Открыть страницу регистрации.",
            "Заполнить обязательные поля.",
            "Подтвердить обязательные согласия.",
            "Нажать кнопку регистрации.",
        ],
        "test_data": [
            "Логин: TestUser",
            "Email: testuser@example.com",
            "Телефон: +7 999 123-45-67",
            "Пароль: TestPass123!",
        ],
        "postcondition": "Создана новая учетная запись пользователя.",
        "expected": "Учетная запись создается, пользователь авторизуется или перенаправляется в предусмотренный сценарий входа, ошибки валидации отсутствуют.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Вход с корректными данными",
        "precondition": "Существует учетная запись пользователя.",
        "steps": [
            "Открыть страницу входа.",
            "Ввести email или логин и пароль.",
            "Нажать кнопку входа.",
        ],
        "test_data": ["Логин: TestUser", "Пароль: TestPass123!"],
        "postcondition": "Пользовательская сессия активна.",
        "expected": "Пользователь успешно входит в систему и получает доступ к личному кабинету.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Вход с неверным паролем",
        "precondition": "Существует учетная запись пользователя.",
        "steps": [
            "Открыть страницу входа.",
            "Ввести корректный логин и неверный пароль.",
            "Нажать кнопку входа.",
        ],
        "test_data": ["Логин: TestUser", "Пароль: WrongPass123!"],
        "postcondition": "Пользователь остается неавторизованным.",
        "expected": "Система отображает сообщение об ошибке и не авторизует пользователя.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Открытие каталога товаров",
        "precondition": "Сайт доступен.",
        "steps": ["Перейти в раздел каталога."],
        "test_data": ["URL: /catalog"],
        "postcondition": "Открыта страница каталога.",
        "expected": "Открывается страница каталога со списком товаров, фильтрами и элементами навигации.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Поиск товара по ключевому слову",
        "precondition": "В каталоге есть товары.",
        "steps": [
            "Открыть каталог.",
            "Ввести поисковый запрос.",
            "Дождаться обновления выдачи.",
        ],
        "test_data": ["Поисковый запрос: hoodie"],
        "postcondition": "Отображена отфильтрованная выдача.",
        "expected": "Отображаются только товары, соответствующие поисковому запросу.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Фильтрация товаров по категории",
        "precondition": "В каталоге есть несколько категорий.",
        "steps": ["Открыть каталог.", "Выбрать категорию в фильтре."],
        "test_data": ["Категория: Худи"],
        "postcondition": "Список товаров отфильтрован.",
        "expected": "Список товаров обновляется в соответствии с выбранной категорией.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Фильтрация товаров по размеру",
        "precondition": "У товаров в каталоге есть варианты размеров.",
        "steps": ["Открыть каталог.", "Выбрать размер в фильтре."],
        "test_data": ["Размер: M"],
        "postcondition": "Показаны товары выбранного размера.",
        "expected": "В выдаче остаются товары с доступными вариантами выбранного размера.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Фильтр только доступных товаров",
        "precondition": "В каталоге есть товары в наличии и вне наличия.",
        "steps": ["Открыть каталог.", "Включить фильтр наличия."],
        "test_data": ["Фильтр: В наличии = Да"],
        "postcondition": "В выдаче остаются только доступные товары.",
        "expected": "В списке отображаются только товары, доступные для покупки.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Открытие карточки товара",
        "precondition": "В каталоге есть опубликованный товар.",
        "steps": ["Открыть каталог.", "Перейти в карточку выбранного товара."],
        "test_data": ["Товар: Akatsuki Hoodie"],
        "postcondition": "Открыта карточка товара.",
        "expected": "Открывается карточка товара с изображениями, описанием, ценой и вариантами.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Выбор варианта размера",
        "precondition": "Открыта карточка товара с несколькими вариантами.",
        "steps": ["Перейти в карточку товара.", "Выбрать размер."],
        "test_data": ["Размер: L"],
        "postcondition": "Активный вариант товара изменен.",
        "expected": "Активный вариант изменяется, состояние выбора отображается в интерфейсе.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Добавление товара в корзину",
        "precondition": "Открыта карточка товара, выбран доступный вариант.",
        "steps": ["Выбрать размер.", "Нажать кнопку добавления в корзину."],
        "test_data": ["Товар: Akatsuki Hoodie, Размер: M, Цвет: Black"],
        "postcondition": "Позиция добавлена в корзину.",
        "expected": "Товар добавляется в корзину, счетчик корзины обновляется.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Добавление товара в избранное",
        "precondition": "Пользователь авторизован.",
        "steps": ["Открыть каталог или карточку товара.", "Нажать кнопку добавления в избранное."],
        "test_data": ["Пользователь: TestUser"],
        "postcondition": "Товар добавлен в избранное.",
        "expected": "Товар появляется в списке избранного, кнопка меняет состояние.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Удаление товара из избранного",
        "precondition": "Пользователь авторизован, товар уже находится в избранном.",
        "steps": ["Открыть список избранного или карточку товара.", "Нажать кнопку удаления из избранного."],
        "test_data": ["Товар в избранном: Akatsuki Hoodie"],
        "postcondition": "Товар удален из избранного.",
        "expected": "Товар удаляется из избранного, состояние кнопки обновляется.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Просмотр содержимого корзины",
        "precondition": "В корзине есть хотя бы один товар.",
        "steps": ["Открыть корзину."],
        "test_data": ["Корзина: Akatsuki Hoodie x1"],
        "postcondition": "Показано текущее содержимое корзины.",
        "expected": "Отображаются добавленные товары, количество, цена и итоговая сумма.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Изменение количества товара в корзине",
        "precondition": "В корзине есть товар.",
        "steps": ["Открыть корзину.", "Увеличить или уменьшить количество позиции."],
        "test_data": ["Товар в корзине: Akatsuki Hoodie, Количество: 1"],
        "postcondition": "Количество товара изменено.",
        "expected": "Количество товара и итоговая сумма пересчитываются корректно.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Удаление товара из корзины",
        "precondition": "В корзине есть товар.",
        "steps": ["Открыть корзину.", "Удалить позицию."],
        "test_data": ["Товар в корзине: Akatsuki Hoodie, Количество: 1"],
        "postcondition": "Позиция удалена из корзины.",
        "expected": "Позиция удаляется из корзины, итоговая сумма обновляется.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Открытие личного кабинета",
        "precondition": "Пользователь авторизован.",
        "steps": ["Перейти в личный кабинет."],
        "test_data": ["Пользователь: TestUser"],
        "postcondition": "Открыт личный кабинет.",
        "expected": "Открывается кабинет с вкладками профиля, адресов, избранного и заказов.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Редактирование контактных данных",
        "precondition": "Пользователь авторизован.",
        "steps": ["Открыть личный кабинет.", "Изменить данные профиля.", "Сохранить изменения."],
        "test_data": ["Имя: Test User, Телефон: +7 999 765-43-21"],
        "postcondition": "Данные профиля обновлены.",
        "expected": "Обновленные данные сохраняются и повторно отображаются после перезагрузки.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Добавление адреса доставки",
        "precondition": "Пользователь авторизован.",
        "steps": ["Открыть вкладку адресов.", "Заполнить форму нового адреса.", "Сохранить адрес."],
        "test_data": [
            "Получатель: Test User",
            "Телефон: +7 999 123-45-67",
            "Город: Москва",
            "Адрес: ул. Тверская, д. 10, кв. 5",
        ],
        "postcondition": "Новый адрес сохранен.",
        "expected": "Новый адрес появляется в списке адресов пользователя.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Редактирование адреса доставки",
        "precondition": "У пользователя есть сохраненный адрес.",
        "steps": ["Открыть вкладку адресов.", "Выбрать редактирование.", "Изменить данные.", "Сохранить."],
        "test_data": [
            "Существующий адрес: Дом",
            "Новый город: Санкт-Петербург",
            "Новый адрес: Невский пр., д. 25",
        ],
        "postcondition": "Адрес обновлен.",
        "expected": "Адрес обновляется и отображается с новыми значениями.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Открытие страницы рекомендаций",
        "precondition": "Сайт доступен.",
        "steps": ["Перейти в раздел рекомендаций."],
        "test_data": ["URL: /fitting"],
        "postcondition": "Открыта страница рекомендаций.",
        "expected": "Открывается страница рекомендаций или теста с доступными шагами сценария.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Прохождение теста рекомендаций",
        "precondition": "Пользователь авторизован.",
        "steps": ["Открыть раздел рекомендаций.", "Последовательно ответить на вопросы.", "Завершить тест."],
        "test_data": [
            "Рост: 178 см",
            "Вес: 74 кг",
            "Предпочитаемый стиль: streetwear",
        ],
        "postcondition": "Результаты теста сохранены.",
        "expected": "Результаты теста сохраняются и доступны для дальнейшего использования в системе.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Открытие страницы оформления заказа",
        "precondition": "В корзине есть товар.",
        "steps": ["Перейти в checkout."],
        "test_data": ["Корзина: Akatsuki Hoodie x1, Tokyo Pants x1"],
        "postcondition": "Открыта страница checkout.",
        "expected": "Открывается страница оформления заказа с данными корзины и формой оформления.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Оформление заказа с корректными данными",
        "precondition": "Пользователь авторизован, в корзине есть товар, настроены доставка и способ оплаты.",
        "steps": [
            "Открыть checkout.",
            "Выбрать адрес доставки.",
            "Выбрать способ доставки.",
            "Выбрать способ оплаты.",
            "Подтвердить обязательные согласия.",
            "Подтвердить оформление.",
        ],
        "test_data": [
            "Адрес: Москва, ул. Тверская, д. 10",
            "Способ доставки: courier",
            "Способ оплаты: card",
        ],
        "postcondition": "Создан новый заказ.",
        "expected": "Заказ создается, отображается успешный переход к сценарию оплаты или подтверждения заказа.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Попытка оформления без обязательных согласий",
        "precondition": "В корзине есть товар.",
        "steps": ["Открыть checkout.", "Не отмечать обязательные согласия.", "Подтвердить оформление."],
        "test_data": ["Корзина: Akatsuki Hoodie x1, согласия не установлены"],
        "postcondition": "Заказ не создается.",
        "expected": "Система не создает заказ и отображает сообщение о необходимости подтверждения согласий.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Возврат со страницы оплаты с успешным статусом",
        "precondition": "Создан заказ и инициирована оплата.",
        "steps": ["Завершить оплату в платежном сценарии.", "Вернуться на сайт."],
        "test_data": ["Заказ: ORD-10025, статус оплаты: succeeded"],
        "postcondition": "Заказ переведен в оплаченный статус.",
        "expected": "Страница возврата показывает успешный статус оплаты, заказ получает актуальный статус.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Возврат со страницы оплаты с ошибкой",
        "precondition": "Создан заказ и инициирована оплата.",
        "steps": ["Прервать оплату или получить неуспешный статус.", "Вернуться на сайт."],
        "test_data": ["Заказ: ORD-10026, статус оплаты: failed"],
        "postcondition": "Заказ остается в неуспешном или ожидающем статусе.",
        "expected": "Пользователю отображается информация об ошибке или неуспешной оплате, повторная оплата остается возможной.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Просмотр истории заказов",
        "precondition": "Пользователь авторизован, у него есть хотя бы один заказ.",
        "steps": ["Открыть личный кабинет.", "Перейти во вкладку заказов."],
        "test_data": ["Пользователь: TestUser, заказы: ORD-10020, ORD-10021"],
        "postcondition": "Открыт список заказов.",
        "expected": "Отображается список оформленных заказов с основными статусами и суммами.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Просмотр деталей заказа и отслеживания",
        "precondition": "У пользователя есть заказ.",
        "steps": ["Открыть список заказов.", "Перейти в карточку конкретного заказа."],
        "test_data": ["Заказ: ORD-10020"],
        "postcondition": "Открыта детальная страница заказа.",
        "expected": "Открывается страница заказа с деталями, статусом и, при наличии, данными отслеживания.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Открытие страницы поиска",
        "precondition": "Сайт доступен.",
        "steps": ["Перейти на страницу поиска."],
        "test_data": ["URL: /search"],
        "postcondition": "Открыта страница поиска.",
        "expected": "Страница поиска открывается корректно, доступны поле ввода и выдача результатов.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Отправка формы обратной связи",
        "precondition": "Сайт доступен.",
        "steps": ["Открыть страницу контактов.", "Заполнить форму.", "Отправить сообщение."],
        "test_data": [
            "Имя: Test User",
            "Email: testuser@example.com",
            "Тема: delivery",
            "Сообщение: Уточните срок доставки заказа ORD-10020",
        ],
        "postcondition": "Заявка обратной связи отправлена.",
        "expected": "Форма успешно отправляется, пользователь получает подтверждение.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Отображение состояния ошибки при недоступности данных",
        "precondition": "Один из API-запросов возвращает ошибку.",
        "steps": ["Открыть страницу, зависящую от API.", "Смоделировать ошибку загрузки."],
        "test_data": ["Сценарий: HTTP 500 при загрузке каталога"],
        "postcondition": "Пользователь видит состояние ошибки.",
        "expected": "Интерфейс показывает понятное сообщение об ошибке и предлагает повторить попытку.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Отображение загрузочных состояний",
        "precondition": "Сайт доступен.",
        "steps": ["Открыть страницу каталога, товара или кабинета при замедленном соединении."],
        "test_data": ["Сценарий: задержка ответа API 3-5 секунд"],
        "postcondition": "Показаны промежуточные состояния загрузки.",
        "expected": "Во время загрузки отображаются skeleton-заглушки или иные промежуточные состояния.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Открытие страницы политики конфиденциальности",
        "precondition": "Сайт доступен.",
        "steps": ["Перейти на страницу политики конфиденциальности."],
        "test_data": ["URL: /privacy"],
        "postcondition": "Открыта страница политики конфиденциальности.",
        "expected": "Страница открывается и отображает содержимое без ошибок.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Открытие страницы оферты и возвратов",
        "precondition": "Сайт доступен.",
        "steps": ["Перейти на страницу оферты.", "Перейти на страницу возвратов."],
        "test_data": ["URL: /offer"],
        "postcondition": "Открыты правовые страницы сайта.",
        "expected": "Обе страницы открываются корректно и доступны для чтения.",
        "actual": "",
        "status": "Не выполнен",
    },
    {
        "title": "Доступ в административную панель",
        "precondition": "Существует учетная запись администратора.",
        "steps": ["Открыть административную панель.", "Авторизоваться как администратор."],
        "test_data": ["Логин: admin, Пароль: AdminPass123!"],
        "postcondition": "Администратор вошел в административную панель.",
        "expected": "Администратор получает доступ к панели управления, обычный пользователь доступ не получает.",
        "actual": "",
        "status": "Не выполнен",
    },
]


def format_step(index: int, text: str) -> str:
    letters = "абвгдежзийклмнопрстуфхцчшщ"
    prefix = letters[index] if index < len(letters) else f"{index + 1}"
    return f"{prefix}) {text}"


def clear_sheet_content(worksheet, start_row: int) -> None:
    merged_ranges = list(worksheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        if merged_range.min_row >= start_row:
            worksheet.unmerge_cells(str(merged_range))
    if worksheet.max_row >= start_row:
        worksheet.delete_rows(start_row, worksheet.max_row - start_row + 1)


def copy_row_style(worksheet, source_row: int, target_row: int) -> None:
    source_height = worksheet.row_dimensions[source_row].height
    if source_height is not None:
        worksheet.row_dimensions[target_row].height = source_height
    for column in range(1, 10):
        source_cell = worksheet.cell(source_row, column)
        target_cell = worksheet.cell(target_row, column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def fill_case(worksheet, start_row: int, case_index: int, case: dict) -> int:
    lines_count = max(len(case["steps"]), len(case["test_data"]), 1)
    for offset in range(lines_count):
        source_row = 3 if offset == 0 else 4
        copy_row_style(worksheet, source_row, start_row + offset)

    end_row = start_row + lines_count - 1

    if lines_count > 1:
        for column in ("B", "C", "F", "G", "H", "I"):
            worksheet.merge_cells(f"{column}{start_row}:{column}{end_row}")

    worksheet[f"B{start_row}"] = f"{case_index}) {case['title']}"
    worksheet[f"C{start_row}"] = case["precondition"]
    worksheet[f"F{start_row}"] = case["postcondition"]
    worksheet[f"G{start_row}"] = case["expected"]
    worksheet[f"H{start_row}"] = case["actual"]
    worksheet[f"I{start_row}"] = case["status"]

    for offset in range(lines_count):
        step_value = case["steps"][offset] if offset < len(case["steps"]) else ""
        data_value = case["test_data"][offset] if offset < len(case["test_data"]) else ""
        if step_value:
            worksheet[f"D{start_row + offset}"] = format_step(offset, step_value)
        if data_value:
            worksheet[f"E{start_row + offset}"] = data_value

    return end_row + 1


def main() -> None:
    project_root = Path(__file__).resolve().parents[1]
    template_path = project_root / "docs" / "template-test-case.xlsx"
    output_path = project_root / "docs" / "manual-test-cases.xlsx"

    workbook = load_workbook(template_path)
    worksheet = workbook.active
    clear_sheet_content(worksheet, start_row=3)

    row_pointer = 3
    for index, case in enumerate(TEST_CASES, start=1):
        row_pointer = fill_case(worksheet, row_pointer, index, case)

    workbook.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
